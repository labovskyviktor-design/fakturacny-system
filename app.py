"""
Fakturačný systém - Flask aplikácia
Klon systémov SuperFaktúra / Kros
"""
import os
import io
import csv
from datetime import date, timedelta
from collections import defaultdict
from flask import Flask, render_template, request, redirect, url_for, flash, make_response, Response, jsonify
import logging
import traceback
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import db, User, Supplier, Client, Invoice, InvoiceItem, ActivityLog, InvoiceView, RecurringInvoice
from utils.company_lookup import lookup_company
from utils.pay_by_square import generate_qr_code_base64, generate_sepa_qr
from utils.email_service import mail
import base64
import sentry_sdk
from sentry_sdk.integrations.flask import FlaskIntegration
from utils import (
    suma_slovom, format_currency, format_date_sk,
    get_payment_method_label, get_status_label, get_status_color,
    generate_pay_by_square
)

# Inicializácia Sentry (ak je DSN dostupné)
if os.environ.get('SENTRY_DSN'):
    sentry_sdk.init(
        dsn=os.environ.get('SENTRY_DSN'),
        integrations=[FlaskIntegration()],
        traces_sample_rate=0.5,
        profiles_sample_rate=0.5,
    )

from config import get_config

# Vytvoríme Flask aplikáciu
app = Flask(__name__)

# Načítanie konfigurácie
app.config.from_object(get_config())

# Inicializácia rozšírení
db.init_app(app)
mail.init_app(app)

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Pre prístup sa musíte prihlásiť.'
login_manager.login_message_category = 'error'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Vytvorenie tabuliek pri štarte
# Vytvorenie tabuliek pri štarte
with app.app_context():
    try:
        db.create_all()
        app.logger.info("Databázové tabuľky boli úspešne inicializované.")
    except Exception as e:
        app.logger.error(f"FATAL ERROR: Nepodarilo sa inicializovať databázu: {e}")
        app.logger.error(traceback.format_exc())

# Registrácia pomocných funkcií do Jinja2
app.jinja_env.globals.update(
    suma_slovom=suma_slovom,
    format_currency=format_currency,
    format_date_sk=format_date_sk,
    get_payment_method_label=get_payment_method_label,
    get_status_label=get_status_label,
    get_status_color=get_status_color,
    generate_pay_by_square=generate_pay_by_square,
)


@app.errorhandler(500)
def internal_error(error):
    """Global error handler pre 500 chyby"""
    # Logovanie chyby
    app.logger.error('Server Error: %s', error)
    app.logger.error(traceback.format_exc())
    
    # Rollback databázovej session
    try:
        db.session.rollback()
    except Exception as e:
        app.logger.error(f'Rollback failed: {e}')
        
    return render_template('500.html'), 500


@app.errorhandler(Exception)
def handle_exception(e):
    """Handler pre všetky neočakávané výnimky"""
    # pass through HTTP errors
    if hasattr(e, "code"):
        return e
        
    app.logger.error(f"Unhandled Exception: {e}")
    app.logger.error(traceback.format_exc())
    
    try:
        db.session.rollback()
    except:
        pass

    return render_template('500.html'), 500



# ==============================================================================
# AUTENTIFIKÁCIA
# ==============================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Prihlásenie používateľa"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        remember = 'remember' in request.form
        
        user = User.query.filter_by(email=email).first()
        
        if user and user.check_password(password):
            login_user(user, remember=remember)
            next_page = request.args.get('next')
            flash(f'Vitajte, {user.name}!', 'success')
            return redirect(next_page or url_for('dashboard'))
        else:
            flash('Nesprávny email alebo heslo.', 'error')
    
    return render_template('auth/login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    """Registrácia nového používateľa"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        company_name = request.form.get('company_name', '').strip()
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        password_confirm = request.form.get('password_confirm', '')
        
        # Validácia
        if not name or not email or not password:
            flash('Vyplňte všetky povinné polia.', 'error')
            return render_template('auth/register.html')
        
        if password != password_confirm:
            flash('Heslá sa nezhodujú.', 'error')
            return render_template('auth/register.html')
        
        if len(password) < 6:
            flash('Heslo musí mať aspoň 6 znakov.', 'error')
            return render_template('auth/register.html')
        
        if User.query.filter_by(email=email).first():
            flash('Používateľ s týmto emailom už existuje.', 'error')
            return render_template('auth/register.html')
        
        # Vytvorenie používateľa
        user = User(name=name, email=email, company_name=company_name)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        
        # Prihlásime používateľa
        login_user(user)
        flash('Registrácia úspešná! Vitajte vo Fakturačnom systéme.', 'success')
        return redirect(url_for('supplier_settings'))  # Presmerujeme na nastavenie firmy
    
    return render_template('auth/register.html')


@app.route('/logout', methods=['GET', 'POST'])
@login_required
def logout():
    """Odhlasenie pouzivatela"""
    # Ak je to demo uzivatel, vymazeme jeho data
    if current_user.email and current_user.email.startswith('demo_'):
        demo_user_id = current_user.id
        logout_user()
        # Vymazeme demo data
        try:
            Invoice.query.filter_by(user_id=demo_user_id).delete()
            Client.query.filter_by(user_id=demo_user_id).delete()
            Supplier.query.filter_by(user_id=demo_user_id).delete()
            ActivityLog.query.filter_by(user_id=demo_user_id).delete()
            User.query.filter_by(id=demo_user_id).delete()
            db.session.commit()
        except:
            db.session.rollback()
        flash('Demo session ukoncena. Data boli vymazane.', 'success')
    else:
        logout_user()
        flash('Boli ste uspesne odhlaseny.', 'success')
    return redirect(url_for('login'))


@app.route('/demo')
def demo_login():
    """Demo prihlasenie bez registracie"""
    import uuid
    
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    # Vytvorime jedinecneho demo uzivatela
    demo_id = str(uuid.uuid4())[:8]
    demo_email = f'demo_{demo_id}@fakturask.demo'
    
    # Vytvorime demo uzivatela
    demo_user = User(
        name='Demo Pouzivatel',
        email=demo_email,
        company_name='Demo Firma s.r.o.'
    )
    demo_user.set_password(demo_id)  # Nahodne heslo
    db.session.add(demo_user)
    db.session.commit()
    
    # Vytvorime demo dodavatela
    demo_supplier = Supplier(
        user_id=demo_user.id,
        name='Demo Firma s.r.o.',
        street='Hlavna 123',
        city='Bratislava',
        zip_code='81101',
        country='Slovenska republika',
        ico='12345678',
        dic='2012345678',
        is_vat_payer=True,
        ic_dph='SK2012345678',
        bank_name='Tatra banka, a.s.',
        iban='SK3111000000002612012345',  # Valid SK IBAN format
        email='demo@firma.sk',
        phone='+421 900 123 456',
        invoice_prefix='DEMO'
    )
    db.session.add(demo_supplier)
    
    # Vytvorime demo klienta
    demo_client = Client(
        user_id=demo_user.id,
        name='Ukazkovy Klient s.r.o.',
        street='Testovacia 456',
        city='Kosice',
        zip_code='04001',
        country='Slovenska republika',
        ico='87654321',
        dic='2087654321',
        email='klient@ukazka.sk'
    )
    db.session.add(demo_client)
    db.session.commit()
    
    # Prihlasime demo uzivatela
    login_user(demo_user)
    
    flash('Vitajte v demo rezime! Mozete si vyskusat vsetky funkcie. Data sa neukladaju a budu vymazane po odhlaseni.', 'warning')
    return redirect(url_for('dashboard'))


# ==============================================================================
# STATICKE STRANKY
# ==============================================================================

@app.route('/terms')
def terms():
    """Podmienky pouzivania"""
    return render_template('terms.html')


@app.route('/gdpr')
def gdpr():
    """Ochrana osobnych udajov (GDPR)"""
    return render_template('gdpr.html')


# ==============================================================================
# DASHBOARD
# ==============================================================================

@app.route('/')
@login_required
def dashboard():
    """Hlavný dashboard s prehľadom a analytics"""
    app.logger.info("Accessing dashboard...")
    try:
        supplier = Supplier.query.filter_by(user_id=current_user.id).first()
        
        # Všetky faktúry tohto používateľa
        app.logger.info("Loading invoices...")
        try:
            invoices = Invoice.query.filter_by(user_id=current_user.id).all()
            app.logger.info(f"Loaded {len(invoices)} invoices")
        except Exception as e:
            app.logger.error(f"Failed to load invoices: {e}")
            raise e
        
        # Aktualizujeme stavy po splatnosti
        app.logger.info("Checking overdue status...")
        try:
            changes_made = False
            for inv in invoices:
                # Store original status
                original_status = inv.status
                inv.check_overdue()
                if inv.status != original_status:
                    changes_made = True
            
            if changes_made:
                db.session.commit()
                app.logger.info("Overdue statuses updated")
        except Exception as e:
            app.logger.error(f"Failed to update overdue status: {e}")
            # Non-critical, continue
        
        # Základné štatistiky
        total_invoices = len(invoices)
        paid_invoices = [i for i in invoices if i.status == Invoice.STATUS_PAID]
        overdue_invoices = [i for i in invoices if i.is_overdue]
        issued_invoices = [i for i in invoices if i.status == Invoice.STATUS_ISSUED]
        
        total_revenue = sum(i.total for i in paid_invoices)
        total_pending = sum(i.total for i in issued_invoices)
        total_overdue = sum(i.total for i in overdue_invoices)
        
        # === ANALYTICS ===
        app.logger.info("Calculating analytics...")
        try:
            active_invoices = [i for i in invoices if i.status != Invoice.STATUS_CANCELLED]
            total_invoiced = sum(i.total for i in active_invoices)
            total_profit = sum(i.profit for i in paid_invoices)
            total_cost = sum(i.total_cost for i in paid_invoices)
            expected_income = total_pending
            
            # Top odberateľ
            client_totals = defaultdict(float)
            for inv in paid_invoices:
                client_totals[inv.client_id] += inv.total
            
            top_client = None
            top_client_amount = 0
            if client_totals:
                top_client_id = max(client_totals, key=client_totals.get)
                top_client = Client.query.filter_by(id=top_client_id, user_id=current_user.id).first()
                top_client_amount = client_totals[top_client_id]
            
            # Mesačný prehľad
            monthly_data = []
            today = date.today()
            for i in range(5, -1, -1):
                month_start = date(today.year, today.month, 1) - timedelta(days=30*i)
                # month_name = month_start.strftime('%b %Y') # Simple string format
                
                # Bezpečný výpočet mesiaca
                month_revenue = sum(
                    inv.total for inv in paid_invoices 
                    if inv.paid_date and inv.paid_date.month == month_start.month and inv.paid_date.year == month_start.year
                )
                month_profit = sum(
                    inv.profit for inv in paid_invoices 
                    if inv.paid_date and inv.paid_date.month == month_start.month and inv.paid_date.year == month_start.year
                )
                
                monthly_data.append({
                    'month': month_start.strftime('%m/%Y'),
                    'revenue': month_revenue,
                    'profit': month_profit
                })
        except Exception as e:
            app.logger.error(f"Analytics calculation failed: {e}")
            app.logger.error(traceback.format_exc())
            # Fallback values
            total_invoiced = 0
            total_profit = 0
            total_cost = 0
            expected_income = 0
            top_client = None
            top_client_amount = 0
            monthly_data = []

        recent_activity = ActivityLog.query.filter_by(user_id=current_user.id).order_by(ActivityLog.created_at.desc()).limit(10).all()
        recent_invoices = Invoice.query.filter_by(user_id=current_user.id).order_by(Invoice.created_at.desc()).limit(10).all()
        
        return render_template('dashboard.html',
            supplier=supplier,
            total_invoices=total_invoices,
            paid_count=len(paid_invoices),
            overdue_count=len(overdue_invoices),
            issued_count=len(issued_invoices),
            total_revenue=total_revenue,
            total_pending=total_pending,
            total_overdue=total_overdue,
            total_invoiced=total_invoiced,
            total_profit=total_profit,
            total_cost=total_cost,
            expected_income=expected_income,
            top_client=top_client,
            top_client_amount=top_client_amount,
            monthly_data=monthly_data,
            recent_activity=recent_activity,
            recent_invoices=recent_invoices
        )
    except Exception as e:
        app.logger.error(f"CRITICAL DASHBOARD ERROR: {e}")
        app.logger.error(traceback.format_exc())
        return render_template('500.html'), 500


# ==============================================================================
# KLIENTI
# ==============================================================================

@app.route('/clients')
@login_required
def clients_list():
    """Zoznam klientov"""
    clients = Client.query.filter_by(user_id=current_user.id).order_by(Client.name).all()
    return render_template('clients.html', clients=clients)


@app.route('/clients/add', methods=['GET', 'POST'])
@login_required
def client_add():
    """Pridanie nového klienta"""
    if request.method == 'POST':
        client = Client(
            user_id=current_user.id,
            name=request.form['name'],
            contact_person=request.form.get('contact_person', ''),
            street=request.form['street'],
            city=request.form['city'],
            zip_code=request.form['zip_code'],
            country=request.form.get('country', 'Slovenská republika'),
            ico=request.form.get('ico', ''),
            dic=request.form.get('dic', ''),
            ic_dph=request.form.get('ic_dph', ''),
            email=request.form.get('email', ''),
            phone=request.form.get('phone', ''),
            note=request.form.get('note', '')
        )
        db.session.add(client)
        db.session.commit()
        flash(f'Klient "{client.name}" bol úspešne pridaný.', 'success')
        return redirect(url_for('clients_list'))
    
    return render_template('client_form.html', client=None)


@app.route('/clients/<int:client_id>/edit', methods=['GET', 'POST'])
@login_required
def client_edit(client_id):
    """Uprava klienta"""
    client = Client.query.filter_by(id=client_id, user_id=current_user.id).first_or_404()
    
    if request.method == 'POST':
        client.name = request.form['name']
        client.contact_person = request.form.get('contact_person', '')
        client.street = request.form['street']
        client.city = request.form['city']
        client.zip_code = request.form['zip_code']
        client.country = request.form.get('country', 'Slovenská republika')
        client.ico = request.form.get('ico', '')
        client.dic = request.form.get('dic', '')
        client.ic_dph = request.form.get('ic_dph', '')
        client.email = request.form.get('email', '')
        client.phone = request.form.get('phone', '')
        client.note = request.form.get('note', '')
        
        db.session.commit()
        flash(f'Klient "{client.name}" bol úspešne upravený.', 'success')
        return redirect(url_for('clients_list'))
    
    return render_template('client_form.html', client=client)


@app.route('/clients/<int:client_id>/delete', methods=['POST'])
@login_required
def client_delete(client_id):
    """Vymazanie klienta"""
    client = Client.query.filter_by(id=client_id, user_id=current_user.id).first_or_404()
    
    # Kontrola či nemá faktúry
    invoice_count = Invoice.query.filter_by(client_id=client_id, user_id=current_user.id).count()
    if invoice_count > 0:
        flash(f'Klient "{client.name}" nemôže byť vymazaný, pretože má faktúry.', 'error')
        return redirect(url_for('clients_list'))
    
    name = client.name
    db.session.delete(client)
    db.session.commit()
    flash(f'Klient "{name}" bol úspešne vymazaný.', 'success')
    return redirect(url_for('clients_list'))


# ==============================================================================
# FAKTÚRY
# ==============================================================================

@app.route('/invoices')
@login_required
def invoices_list():
    """Zoznam faktúr"""
    app.logger.info("Accessing invoices list...")
    try:
        status_filter = request.args.get('status', '')
        search_query = request.args.get('q', '')
        
        query = Invoice.query.filter_by(user_id=current_user.id)
        
        # Aktualizujeme stavy pred filtrovaním
        try:
            all_invoices = Invoice.query.filter_by(user_id=current_user.id).all()
            changes = False
            for inv in all_invoices:
                original = inv.status
                inv.check_overdue()
                if inv.status != original:
                    changes = True
            if changes:
                db.session.commit()
                app.logger.info("Overdue statuses updated in list view")
        except Exception as e:
            app.logger.error(f"Failed to update overdue status in list: {e}")
            # Continue anyway
        
        # Filtre
        if status_filter == 'overdue':
            query = query.filter(Invoice.status == Invoice.STATUS_OVERDUE)
        elif status_filter:
            query = query.filter_by(status=status_filter)
        
        # Vyhľadávanie
        if search_query:
            query = query.join(Client).filter(
                (Invoice.invoice_number.contains(search_query)) |
                (Invoice.variable_symbol.contains(search_query)) |
                (Client.name.contains(search_query))
            )
        
        invoices = query.order_by(Invoice.created_at.desc()).all()
        app.logger.info(f"Loaded {len(invoices)} invoices for list")
        
        return render_template('invoices.html', 
            invoices=invoices,
            status_filter=status_filter,
            search_query=search_query
        )
    except Exception as e:
        app.logger.error(f"CRITICAL INVOICE LIST ERROR: {e}")
        app.logger.error(traceback.format_exc())
        return render_template('500.html'), 500


@app.route('/invoices/add', methods=['GET', 'POST'])
@login_required
def invoice_add():
    """Vytvorenie novej faktúry"""
    supplier = Supplier.query.filter_by(user_id=current_user.id).first()
    clients = Client.query.filter_by(user_id=current_user.id).order_by(Client.name).all()
    
    if not supplier:
        flash('Najprv musíte nastaviť údaje dodávateľa.', 'error')
        return redirect(url_for('supplier_settings'))
    
    if not clients:
        flash('Najprv musíte pridať aspoň jedného klienta.', 'error')
        return redirect(url_for('client_add'))
    
    if request.method == 'POST':
        app.logger.info("Processing invoice creation...")
        try:
            # Získame klienta
            client_id = request.form.get('client_id')
            if not client_id:
                flash('Vyberte klienta.', 'error')
                return redirect(url_for('invoice_add'))
            
            client = Client.query.filter_by(id=client_id, user_id=current_user.id).first()
            if not client:
                flash('Klient neexistuje.', 'error')
                return redirect(url_for('invoice_add'))
            
            # Generujeme číslo faktúry
            invoice_number = supplier.get_next_invoice_number()
            app.logger.info(f"Generated invoice number: {invoice_number}")
            
            # Vytvoríme faktúru
            invoice = Invoice(
                user_id=current_user.id,
                invoice_number=invoice_number,
                variable_symbol=invoice_number.replace('/', ''),
                supplier_id=supplier.id,
                client_id=client.id,
                issue_date=date.fromisoformat(request.form['issue_date']),
                delivery_date=date.fromisoformat(request.form['delivery_date']),
                due_date=date.fromisoformat(request.form['due_date']),
                payment_method=request.form.get('payment_method', 'prevod'),
                vat_rate=float(request.form.get('vat_rate', 0) or 0),
                note=request.form.get('note', ''),
                status=Invoice.STATUS_ISSUED
            )
            
            db.session.add(invoice)
            db.session.flush()
            app.logger.info(f"Invoice {invoice.id} flushed to DB")
            
            # Pridáme položky
            descriptions = request.form.getlist('item_description[]')
            item_notes = request.form.getlist('item_note[]')
            quantities = request.form.getlist('item_quantity[]')
            units = request.form.getlist('item_unit[]')
            unit_prices = request.form.getlist('item_unit_price[]')
            cost_prices = request.form.getlist('item_cost_price[]')
            
            items_added = 0
            for i, desc in enumerate(descriptions):
                if desc and desc.strip():
                    try:
                        qty = float(quantities[i]) if i < len(quantities) and quantities[i] else 1
                        price = float(unit_prices[i]) if i < len(unit_prices) and unit_prices[i] else 0
                        cost = float(cost_prices[i]) if i < len(cost_prices) and cost_prices[i] else 0
                    except (ValueError, IndexError):
                        qty, price, cost = 1, 0, 0
                    
                    item = InvoiceItem(
                        invoice_id=invoice.id,
                        description=desc.strip(),
                        item_note=item_notes[i] if i < len(item_notes) else '',
                        quantity=qty,
                        unit=units[i] if i < len(units) and units[i] else 'ks',
                        unit_price=price,
                        cost_price=cost,
                        position=i
                    )
                    item.calculate_total()
                    db.session.add(item)
                    items_added += 1
            
            if items_added == 0:
                db.session.rollback()
                flash('Pridajte aspoň jednu položku.', 'error')
                return redirect(url_for('invoice_add'))
            
            db.session.flush()
            invoice.calculate_totals()
            
            # Activity log
            app.logger.info("Logging activity...")
            ActivityLog.log(
                ActivityLog.ACTION_INVOICE_CREATED,
                f'Faktúra {invoice.invoice_number} vytvorená pre {client.name}',
                user_id=current_user.id,
                invoice_id=invoice.id,
                client_id=client.id,
                extra_data={'total': invoice.total}
            )
            
            db.session.commit()
            app.logger.info("Invoice committed successfully.")
            
            flash(f'Faktúra {invoice.invoice_number} bola úspešne vytvorená.', 'success')
            return redirect(url_for('invoice_detail', invoice_id=invoice.id))
            
        except ValueError as e:
            db.session.rollback()
            print(f'ValueError pri vytváraní faktúry: {e}')
            import traceback
            traceback.print_exc()
            flash(f'Chyba vo formáte dát: {str(e)}', 'error')
            return redirect(url_for('invoice_add'))
            
        except Exception as e:
            db.session.rollback()
            print(f'Chyba pri vytváraní faktúry: {e}')
            import traceback
            traceback.print_exc()
            flash(f'Chyba pri vytváraní faktúry: {str(e)}', 'error')
            return redirect(url_for('invoice_add'))
    
    # Predvyplnené dátumy
    today = date.today()
    default_due_date = today + timedelta(days=14)
    
    return render_template('invoice_form.html',
        invoice=None,
        supplier=supplier,
        clients=clients,
        today=today,
        default_due_date=default_due_date
    )


@app.route('/invoices/<int:invoice_id>')
@login_required
def invoice_detail(invoice_id):
    """Detail faktúry"""
    app.logger.info(f"Viewing invoice {invoice_id}")
    try:
        invoice = Invoice.query.filter_by(id=invoice_id, user_id=current_user.id).first_or_404()
        app.logger.info(f"Invoice loaded: {invoice.invoice_number}")
        
        # Generujeme QR kód
        qr_code = None
        if app.config.get('ENABLE_QR_CODES') and invoice.payment_method == 'prevod' and invoice.supplier.iban:
            app.logger.info("Generating QR code...")
            try:
                qr_code = generate_qr_code_base64(
                    amount=invoice.total,
                    iban=invoice.supplier.iban,
                    swift=invoice.supplier.swift or '',
                    variable_symbol=invoice.variable_symbol,
                    beneficiary_name=invoice.supplier.name,
                    due_date=invoice.due_date.strftime('%Y%m%d')
                )
                app.logger.info("QR code generated successfully")
            except Exception as e:
                app.logger.error(f"Chyba pri generovaní QR kódu: {e}")
                app.logger.error(traceback.format_exc())
        
        # Parametre pre Gmail (Compose link)
        gmail_link = None
        if invoice.client.email:
            from urllib.parse import quote
            subject = quote(f"Faktúra č. {invoice.invoice_number}")
            body = quote(f"Dobrý deň,\n\nv prílohe Vám zasielame faktúru č. {invoice.invoice_number}.\n\nS pozdravom,\n{invoice.supplier.name}")
            gmail_link = f"https://mail.google.com/mail/?view=cm&fs=1&to={invoice.client.email}&su={subject}&body={body}"

        app.logger.info("Rendering template invoice_detail.html")
        return render_template('invoice_detail.html',
            invoice=invoice,
            qr_code=qr_code,
            gmail_link=gmail_link
        )
    except Exception as e:
        app.logger.error(f"CRITICAL ERROR in invoice_detail: {e}")
        app.logger.error(traceback.format_exc())
        raise e  # Let the global handler handle it


def _get_invoice_pdf_data(invoice):
    """Pomocná funkcia na generovanie PDF dát faktúry"""
    # Generujeme QR kód
    qr_code = None
    if invoice.payment_method == 'prevod' and invoice.supplier.iban:
        try:
            qr_code = generate_qr_code_base64(
                amount=invoice.total,
                iban=invoice.supplier.iban,
                swift=invoice.supplier.swift or '',
                variable_symbol=invoice.variable_symbol,
                beneficiary_name=invoice.supplier.name,
                due_date=invoice.due_date.strftime('%Y%m%d')
            )
        except Exception as e:
            app.logger.error(f"Chyba pri generovaní QR kódu pre PDF: {e}")
    
    # Renderujeme HTML šablónu
    html = render_template('invoice_pdf.html',
        invoice=invoice,
        qr_code=qr_code
    )
    
    # Generovanie PDF cez pdfkit (wkhtmltopdf) - EXTRÉMNE ROBUSTNÉ VYHĽADÁVANIE
    try:
        import pdfkit
        import os
        import shutil
        import subprocess
        
        wk_path = None
        xvfb_prefix = ""
        
        # 1. HLADANIE BINARKY (Viacúrovňové)
        # Pokus A: Naša lokálna binárka zo symlinku v builde
        local_bin = os.path.join(os.getcwd(), 'bin', 'wkhtmltopdf')
        local_xvfb = os.path.join(os.getcwd(), 'bin', 'xvfb-run')
        
        if os.path.exists(local_bin):
            wk_path = local_bin
            if os.path.exists(local_xvfb):
                xvfb_prefix = f"{local_xvfb} -a "
        else:
            # Pokus B: Systémový PATH
            wk_path = shutil.which('wkhtmltopdf')
            xvfb_path = shutil.which('xvfb-run')
            if xvfb_path:
                xvfb_prefix = f"{xvfb_path} -a "

        # Pokus C: Natvrdo Railway Nix store (ak vieme cestu alebo find)
        if not wk_path:
            try:
                res = subprocess.run(['find', '/nix/store', '-name', 'wkhtmltopdf', '-type', 'f', '-executable'], 
                                   capture_output=True, text=True, timeout=5)
                if res.stdout.strip(): wk_path = res.stdout.strip().splitlines()[0]
            except: pass

        if not wk_path:
            raise Exception("wkhtmltopdf binary not found. Path searched including bin/ and PATH.")

        # 2. KONFIGURÁCIA
        app.logger.info(f"Using PDF Engine: {xvfb_prefix}{wk_path}")
        config = pdfkit.configuration(wkhtmltopdf=f"{xvfb_prefix}{wk_path}")

        options = {
            'page-size': 'A4',
            'orientation': 'Portrait',
            'margin-top': '0',
            'margin-right': '0',
            'margin-bottom': '0',
            'margin-left': '0',
            'encoding': "UTF-8",
            'no-outline': None,
            'quiet': ''
        }
        
        # 3. GENEROVANIE
        pdf_data = pdfkit.from_string(html, False, configuration=config, options=options)
        return pdf_data, "application/pdf", True
                
    except Exception as e:
        app.logger.error(f"PDF generation failed: {str(e)}")
        # Ak zlyhá všetko, vrátime aspoň HTML verziu s oznamom o chybe
        return html.encode('utf-8'), "text/html", f"ERROR: PDF generation failed: {str(e)}"

@app.route('/debug/pdf-test')
@login_required
def debug_pdf_test():
    """Testovacia cesta na overenie funkčnosti pdfkit na serveri"""
    if not current_user.is_authenticated:
        return "Not authorized", 403
        
    try:
        import os
        import shutil
        import subprocess
        
        local_bin = os.path.join(os.getcwd(), 'bin', 'wkhtmltopdf')
        local_xvfb = os.path.join(os.getcwd(), 'bin', 'xvfb-run')
        
        results = {
            "Local bin/ folder exists": os.path.exists('bin'),
            "Local wkhtmltopdf exists": os.path.exists(local_bin),
            "Local xvfb-run exists": os.path.exists(local_xvfb),
            "System wkhtmltopdf (which)": shutil.which('wkhtmltopdf'),
            "System xvfb-run (which)": shutil.which('xvfb-run'),
            "Current CWD": os.getcwd()
        }
        
        html_diag = f"""
        <html>
            <body style="font-family: sans-serif; padding: 40px; line-height: 1.6;">
                <h1 style="color: #1e40af;">PDF Binary Diagnostics (Final) 🛠️</h1>
                <ul>
                    {"".join([f"<li><b>{k}:</b> {v}</li>" for k,v in results.items()])}
                </ul>
                <hr>
                <p>Ak vidíte "Local wkhtmltopdf exists: True", tak PDF už na 100% pôjde.</p>
                <p>Ak sú všetky False, skúsime nájsť binárku v /nix/store cez find...</p>
            </body>
        </html>
        """
        return html_diag

    except Exception as e:
        import traceback
        return f"<h1>Debug Failure</h1><pre>{traceback.format_exc()}</pre>", 500

    except Exception as e:
        import traceback
        return f"<h1>Debug Failure</h1><pre>{traceback.format_exc()}</pre>", 500

    except Exception as e:
        import traceback
        return f"<h1>Debug Failure</h1><pre>{traceback.format_exc()}</pre>", 500


@app.route('/invoices/<int:invoice_id>/send', methods=['POST'])
@login_required
def invoice_send_email(invoice_id):
    """Odoslanie faktúry emailom"""
    invoice = Invoice.query.filter_by(id=invoice_id, user_id=current_user.id).first_or_404()
    
    if not invoice.client.email:
        flash('Klient nemá nastavený email.', 'error')
        return redirect(url_for('invoice_detail', invoice_id=invoice.id))
        
    subject = f'Faktúra {invoice.invoice_number} - {invoice.supplier.name}'
    body = f"""Dobrý deň,
    
v prílohe Vám posielame faktúru č. {invoice.invoice_number} na sumu {invoice.total} EUR.

Dátum splatnosti: {invoice.due_date.strftime('%d.%m.%Y')}
Variabilný symbol: {invoice.variable_symbol}

S pozdravom,
{invoice.supplier.name}
"""
    
    # Generujeme prílohu
    pdf_data, content_type, result = _get_invoice_pdf_data(invoice)
    
    if result is True:
        ext = "pdf"
        attachments = [(f"faktura_{invoice.invoice_number}.pdf", content_type, pdf_data)]
    else:
        # Fallback na HTML
        ext = "html"
        error_msg = result if isinstance(result, str) else "Neznáma chyba"
        flash(f"Generovanie PDF zlyhalo: {error_msg}. Súbor bol odoslaný ako HTML.", "warning")
        attachments = [(f"faktura_{invoice.invoice_number}.html", content_type, pdf_data)]
    
    from utils.email_service import send_email
    
    if send_email(subject, invoice.client.email, body, attachments=attachments):
        flash(f'Faktúra bola odoslaná na {invoice.client.email}', 'success')
    else:
        app.logger.error("Failed to send email - check logs")
        flash('Nepodarilo sa odoslať email. Skontrolujte nastavenia (API kľúč).', 'error')
        
    return redirect(url_for('invoice_detail', invoice_id=invoice.id))


@app.route('/invoices/<int:invoice_id>/pdf')
@login_required
def invoice_pdf(invoice_id):
    """Stiahnutie faktúry ako PDF"""
    invoice = Invoice.query.filter_by(id=invoice_id, user_id=current_user.id).first_or_404()
    
    pdf_data, content_type, result = _get_invoice_pdf_data(invoice)
    
    if result is True:
        ext = "pdf"
    else:
        # Fallback na HTML
        ext = "html"
        error_msg = result if isinstance(result, str) else "Neznáma chyba"
        flash(f"Generovanie PDF zlyhalo: {error_msg}. Stiahnutá HTML verzia.", 'warning')
        
    response = make_response(pdf_data)
    response.headers['Content-Type'] = content_type
    response.headers['Content-Disposition'] = f'attachment; filename=faktura_{invoice.invoice_number}.{ext}'
    return response


@app.route('/invoices/<int:invoice_id>/mark-paid', methods=['POST'])
@login_required
def invoice_mark_paid(invoice_id):
    """Označiť faktúru ako uhradenú"""
    invoice = Invoice.query.filter_by(id=invoice_id, user_id=current_user.id).first_or_404()
    invoice.status = Invoice.STATUS_PAID
    invoice.paid_date = date.today()
    
    ActivityLog.log(
        ActivityLog.ACTION_INVOICE_PAID,
        f'Faktúra {invoice.invoice_number} označená ako uhradená',
        user_id=current_user.id,
        invoice_id=invoice.id,
        client_id=invoice.client_id,
        extra_data={'total': invoice.total, 'paid_date': str(invoice.paid_date)}
    )
    
    db.session.commit()
    flash(f'Faktúra {invoice.invoice_number} bola označená ako uhradená.', 'success')
    return redirect(url_for('invoice_detail', invoice_id=invoice.id))


@app.route('/invoices/<int:invoice_id>/cancel', methods=['POST'])
@login_required
def invoice_cancel(invoice_id):
    """Stornovať faktúru"""
    invoice = Invoice.query.filter_by(id=invoice_id, user_id=current_user.id).first_or_404()
    invoice.status = Invoice.STATUS_CANCELLED
    
    ActivityLog.log(
        ActivityLog.ACTION_INVOICE_CANCELLED,
        f'Faktúra {invoice.invoice_number} stornovaná',
        user_id=current_user.id,
        invoice_id=invoice.id,
        client_id=invoice.client_id
    )
    
    db.session.commit()
    flash(f'Faktúra {invoice.invoice_number} bola stornovaná.', 'success')
    return redirect(url_for('invoice_detail', invoice_id=invoice.id))


@app.route('/invoices/<int:invoice_id>/delete', methods=['POST'])
@login_required
def invoice_delete(invoice_id):
    """Vymazať faktúru"""
    invoice = Invoice.query.filter_by(id=invoice_id, user_id=current_user.id).first_or_404()
    number = invoice.invoice_number
    client_id = invoice.client_id
    
    ActivityLog.log(
        ActivityLog.ACTION_INVOICE_DELETED,
        f'Faktúra {number} vymazaná',
        user_id=current_user.id,
        client_id=client_id
    )
    
    db.session.delete(invoice)
    db.session.commit()
    flash(f'Faktúra {number} bola vymazaná.', 'success')
    return redirect(url_for('invoices_list'))


@app.route('/invoices/<int:invoice_id>/edit', methods=['GET', 'POST'])
@login_required
def invoice_edit(invoice_id):
    """Editácia existujúcej faktúry"""
    invoice = Invoice.query.filter_by(id=invoice_id, user_id=current_user.id).first_or_404()
    supplier = Supplier.query.filter_by(user_id=current_user.id).first()
    clients = Client.query.filter_by(user_id=current_user.id).order_by(Client.name).all()
    
    if request.method == 'POST':
        # Aktualizujeme základné údaje
        invoice.client_id = int(request.form['client_id'])
        invoice.issue_date = date.fromisoformat(request.form['issue_date'])
        invoice.delivery_date = date.fromisoformat(request.form['delivery_date'])
        invoice.due_date = date.fromisoformat(request.form['due_date'])
        invoice.payment_method = request.form['payment_method']
        invoice.vat_rate = float(request.form.get('vat_rate', 0))
        invoice.note = request.form.get('note', '')
        invoice.internal_note = request.form.get('internal_note', '')
        
        # Vymažeme staré položky
        for item in invoice.items:
            db.session.delete(item)
        
        # Pridáme nové položky
        descriptions = request.form.getlist('item_description[]')
        item_notes = request.form.getlist('item_note[]')
        quantities = request.form.getlist('item_quantity[]')
        units = request.form.getlist('item_unit[]')
        unit_prices = request.form.getlist('item_unit_price[]')
        cost_prices = request.form.getlist('item_cost_price[]')
        
        for i, desc in enumerate(descriptions):
            if desc.strip():
                item = InvoiceItem(
                    invoice_id=invoice.id,
                    description=desc,
                    item_note=item_notes[i] if i < len(item_notes) else '',
                    quantity=float(quantities[i]) if quantities[i] else 1,
                    unit=units[i] if units[i] else 'ks',
                    unit_price=float(unit_prices[i]) if unit_prices[i] else 0,
                    cost_price=float(cost_prices[i]) if i < len(cost_prices) and cost_prices[i] else 0,
                    position=i
                )
                item.calculate_total()
                db.session.add(item)
        
        db.session.flush()
        invoice.calculate_totals()
        
        ActivityLog.log(
            ActivityLog.ACTION_INVOICE_EDITED,
            f'Faktúra {invoice.invoice_number} upravená',
            user_id=current_user.id,
            invoice_id=invoice.id,
            client_id=invoice.client_id,
            extra_data={'total': invoice.total}
        )
        
        db.session.commit()
        flash(f'Faktúra {invoice.invoice_number} bola aktualizovaná.', 'success')
        return redirect(url_for('invoice_detail', invoice_id=invoice.id))
    
    return render_template('invoice_form.html',
        invoice=invoice,
        supplier=supplier,
        clients=clients,
        today=invoice.issue_date,
        default_due_date=invoice.due_date
    )


@app.route('/invoices/<int:invoice_id>/clone', methods=['POST'])
@login_required
def invoice_clone(invoice_id):
    """Klonovanie faktúry - vytvorí novú faktúru s rovnakými údajmi"""
    original = Invoice.query.filter_by(id=invoice_id, user_id=current_user.id).first_or_404()
    supplier = Supplier.query.filter_by(user_id=current_user.id).first()
    
    # Generujeme nové číslo faktúry
    invoice_number = supplier.get_next_invoice_number()
    today = date.today()
    
    # Vytvoríme novú faktúru
    new_invoice = Invoice(
        invoice_number=invoice_number,
        variable_symbol=invoice_number.replace('FV', '').replace('/', ''),
        user_id=current_user.id,
        supplier_id=supplier.id,
        client_id=original.client_id,
        issue_date=today,
        delivery_date=today,
        due_date=today + timedelta(days=14),
        payment_method=original.payment_method,
        vat_rate=original.vat_rate,
        note=original.note,
        status=Invoice.STATUS_ISSUED
    )
    
    db.session.add(new_invoice)
    db.session.flush()
    
    # Skopírujeme položky
    for orig_item in original.items:
        new_item = InvoiceItem(
            invoice_id=new_invoice.id,
            description=orig_item.description,
            item_note=orig_item.item_note,
            quantity=orig_item.quantity,
            unit=orig_item.unit,
            unit_price=orig_item.unit_price,
            cost_price=orig_item.cost_price,
            position=orig_item.position
        )
        new_item.calculate_total()
        db.session.add(new_item)
    
    db.session.flush()
    new_invoice.calculate_totals()
    
    ActivityLog.log(
        ActivityLog.ACTION_INVOICE_CREATED,
        f'Faktúra {new_invoice.invoice_number} vytvorena klonovaním z {original.invoice_number}',
        user_id=current_user.id,
        invoice_id=new_invoice.id,
        client_id=new_invoice.client_id,
        extra_data={'cloned_from': original.invoice_number}
    )
    
    db.session.commit()
    flash(f'Faktúra {new_invoice.invoice_number} bola vytvorená klonovaním.', 'success')
    return redirect(url_for('invoice_edit', invoice_id=new_invoice.id))


@app.route('/invoices/<int:invoice_id>/internal-note', methods=['POST'])
@login_required
def invoice_internal_note(invoice_id):
    """Aktualizovať internú poznámku faktúry"""
    invoice = Invoice.query.filter_by(id=invoice_id, user_id=current_user.id).first_or_404()
    invoice.internal_note = request.form.get('internal_note', '')
    db.session.commit()
    flash('Interná poznámka bola uložená.', 'success')
    return redirect(url_for('invoice_detail', invoice_id=invoice.id))


@app.route('/invoices/export/csv')
@login_required
def invoices_export_csv():
    """Export faktúr do CSV"""
    invoices = Invoice.query.filter_by(user_id=current_user.id).order_by(Invoice.created_at.desc()).all()
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    
    # Hlavička
    writer.writerow([
        'Číslo faktúry', 'Varia. symbol', 'Klient', 'IČO klienta',
        'Dátum vystavenia', 'Dátum splatnosti', 'Dátum úhrady',
        'Medzisúčet', 'DPH', 'Celkom', 'Stav', 'Forma úhrady'
    ])
    
    # Dáta
    for inv in invoices:
        writer.writerow([
            inv.invoice_number,
            inv.variable_symbol,
            inv.client.name,
            inv.client.ico or '',
            inv.issue_date.strftime('%d.%m.%Y'),
            inv.due_date.strftime('%d.%m.%Y'),
            inv.paid_date.strftime('%d.%m.%Y') if inv.paid_date else '',
            f"{inv.subtotal:.2f}".replace('.', ','),
            f"{inv.vat_amount:.2f}".replace('.', ','),
            f"{inv.total:.2f}".replace('.', ','),
            get_status_label(inv.status),
            get_payment_method_label(inv.payment_method)
        ])
    
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Disposition': f'attachment; filename=faktury_export_{date.today().strftime("%Y%m%d")}.csv',
            'Content-Type': 'text/csv; charset=utf-8'
        }
    )


# ==============================================================================
# NASTAVENIA DODÁVATEĽA
# ==============================================================================

@app.route('/settings', methods=['GET', 'POST'])
@login_required
def supplier_settings():
    """Nastavenia údajov dodávateľa"""
    supplier = Supplier.query.filter_by(user_id=current_user.id).first()
    
    if request.method == 'POST':
        if supplier:
            # Aktualizácia
            supplier.name = request.form['name']
            supplier.street = request.form['street']
            supplier.city = request.form['city']
            supplier.zip_code = request.form['zip_code']
            supplier.country = request.form.get('country', 'Slovenská republika')
            supplier.ico = request.form['ico']
            supplier.dic = request.form.get('dic', '')
            supplier.ic_dph = request.form.get('ic_dph', '')
            supplier.is_vat_payer = 'is_vat_payer' in request.form
            supplier.bank_name = request.form.get('bank_name', '')
            supplier.iban = request.form.get('iban', '')
            supplier.swift = request.form.get('swift', '')
            supplier.email = request.form.get('email', '')
            supplier.phone = request.form.get('phone', '')
            supplier.web = request.form.get('web', '')
            supplier.invoice_prefix = request.form.get('invoice_prefix', '')
        else:
            # Vytvorenie nového
            supplier = Supplier(
                user_id=current_user.id,
                name=request.form['name'],
                street=request.form['street'],
                city=request.form['city'],
                zip_code=request.form['zip_code'],
                country=request.form.get('country', 'Slovenská republika'),
                ico=request.form['ico'],
                dic=request.form.get('dic', ''),
                ic_dph=request.form.get('ic_dph', ''),
                is_vat_payer='is_vat_payer' in request.form,
                bank_name=request.form.get('bank_name', ''),
                iban=request.form.get('iban', ''),
                swift=request.form.get('swift', ''),
                email=request.form.get('email', ''),
                phone=request.form.get('phone', ''),
                web=request.form.get('web', ''),
                invoice_prefix=request.form.get('invoice_prefix', '')
            )
            db.session.add(supplier)
        
        db.session.commit()
        flash('Nastavenia boli uložené.', 'success')
        return redirect(url_for('supplier_settings'))
    
    return render_template('settings.html', supplier=supplier)


# ==============================================================================
# API ENDPOINTS
# ==============================================================================

@app.route('/api/clients/create', methods=['POST'])
@login_required
def api_client_create():
    """
    API endpoint pre vytvorenie nového klienta cez AJAX.
    Používa sa z formulára faktúry.
    """
    try:
        data = request.get_json()
        
        if not data.get('name') or not data.get('street') or not data.get('city') or not data.get('zip_code'):
            return jsonify({'success': False, 'error': 'Vyplňte všetky povinné polia'}), 400
        
        client = Client(
            user_id=current_user.id,
            name=data.get('name', ''),
            contact_person=data.get('contact_person', ''),
            street=data.get('street', ''),
            city=data.get('city', ''),
            zip_code=data.get('zip_code', ''),
            country=data.get('country', 'Slovenská republika'),
            ico=data.get('ico', ''),
            dic=data.get('dic', ''),
            ic_dph=data.get('ic_dph', ''),
            email=data.get('email', ''),
            phone=data.get('phone', '')
        )
        
        db.session.add(client)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'client': {
                'id': client.id,
                'name': client.name,
                'city': client.city
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/rpo/lookup/<ico>')
@login_required
def rpo_lookup(ico):
    """
    API endpoint pre vyhľadanie firmy v RPO podľa IČO.
    Používa sa cez AJAX z formulárov.
    """
    if not ico or len(ico.strip()) < 3:
        return jsonify({'error': 'IČO musí mať aspoň 3 znaky'}), 400
    
    result = lookup_company(ico)
    
    if result:
        return jsonify({
            'success': True,
            'data': result
        })
    else:
        return jsonify({
            'success': False,
            'error': 'Firma s týmto IČO nebola nájdená'
        }), 404


@app.route('/api/upload-stamp', methods=['POST'])
@login_required
def upload_stamp():
    """Nahrávanie pečiatky alebo podpisu"""
    supplier = Supplier.query.filter_by(user_id=current_user.id).first()
    if not supplier:
        return jsonify({'success': False, 'error': 'Najprv nastavťe údaje dodávateľa'}), 400
    
    image_type = request.form.get('type', 'stamp')  # 'stamp' alebo 'signature'
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Nebol nahraný žiaden súbor'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'Nebol vybraný súbor'}), 400
    
    # Skontrolujeme typ súboru
    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif'}
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    
    if ext not in allowed_extensions:
        return jsonify({'success': False, 'error': 'Povolené sú len obrázky (PNG, JPG, GIF)'}), 400
    
    # Konvertujeme na Base64
    file_data = file.read()
    b64_data = base64.b64encode(file_data).decode('utf-8')
    mime_type = f'image/{ext}' if ext != 'jpg' else 'image/jpeg'
    data_uri = f'data:{mime_type};base64,{b64_data}'
    
    # Uložíme
    if image_type == 'signature':
        supplier.signature_image = data_uri
    else:
        supplier.stamp_image = data_uri
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'image_url': data_uri,
        'message': 'Obrázok bol nahraný'
    })


@app.route('/api/remove-stamp', methods=['POST'])
@login_required
def remove_stamp():
    """Odstranenie peciatky alebo podpisu"""
    supplier = Supplier.query.filter_by(user_id=current_user.id).first()
    if not supplier:
        return jsonify({'success': False, 'error': 'Dodavatel neexistuje'}), 400
    
    image_type = request.json.get('type', 'stamp')
    
    if image_type == 'signature':
        supplier.signature_image = None
    else:
        supplier.stamp_image = None
    
    db.session.commit()
    
    return jsonify({'success': True})


# ==============================================================================
# KLIENTSKA ZONA - PUBLIC LINKS
# ==============================================================================

@app.route('/invoice/view/<token>')
def public_invoice_view(token):
    """Verejna stranka faktury pre klienta - bez prihlasenia"""
    from datetime import datetime
    
    invoice = Invoice.query.filter_by(public_token=token).first_or_404()
    
    # Zaznamenaj zobrazenie
    invoice.view_count = (invoice.view_count or 0) + 1
    invoice.last_viewed_at = datetime.utcnow()
    if not invoice.first_viewed_at:
        invoice.first_viewed_at = datetime.utcnow()
    
    # Uloz detailny zaznam
    view_record = InvoiceView(
        invoice_id=invoice.id,
        ip_address=request.remote_addr,
        user_agent=request.user_agent.string[:500] if request.user_agent else None
    )
    db.session.add(view_record)
    db.session.commit()
    
    # Generuj QR kod
    qr_code = None
    if invoice.payment_method == 'prevod' and invoice.supplier.iban:
        try:
            qr_code = generate_qr_code_base64(
                amount=invoice.total,
                iban=invoice.supplier.iban,
                swift=invoice.supplier.swift or '',
                variable_symbol=invoice.variable_symbol,
                beneficiary_name=invoice.supplier.name,
                due_date=invoice.due_date.strftime('%Y%m%d')
            )
        except:
            pass
    
    return render_template('invoice_public.html', invoice=invoice, qr_code=qr_code)


@app.route('/invoices/<int:invoice_id>/generate-link', methods=['POST'])
@login_required
def invoice_generate_link(invoice_id):
    """Vygeneruje verejny link pre fakturu"""
    import secrets
    
    invoice = Invoice.query.filter_by(id=invoice_id, user_id=current_user.id).first_or_404()
    
    if not invoice.public_token:
        invoice.public_token = secrets.token_urlsafe(32)
        db.session.commit()
    
    public_url = url_for('public_invoice_view', token=invoice.public_token, _external=True)
    
    return jsonify({
        'success': True,
        'public_url': public_url,
        'token': invoice.public_token
    })


# ==============================================================================
# DANOVY KALENDAR
# ==============================================================================

@app.route('/tax-calendar')
@login_required
def tax_calendar():
    """Danovy kalendar SR"""
    return render_template('tax_calendar.html')


# ==============================================================================
# EXPORTY PRE UCTOVNICKU
# ==============================================================================

@app.route('/invoices/export/excel')
@login_required
def invoices_export_excel():
    """Export faktur do Excel (.xlsx)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        # Ak nie je openpyxl, pouzijeme CSV format s .xls priponou
        flash('Pre Excel export nainstalujte openpyxl: pip install openpyxl', 'warning')
        return redirect(url_for('invoices_export_csv'))
    
    invoices = Invoice.query.filter_by(user_id=current_user.id).order_by(Invoice.issue_date.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Faktury"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Hlavicka
    headers = ['Cislo faktury', 'Var. symbol', 'Klient', 'ICO', 'Datum vystavenia', 
               'Datum splatnosti', 'Datum uhrady', 'Zaklad DPH', 'DPH', 'Celkom', 'Stav']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data
    for row, inv in enumerate(invoices, 2):
        data = [
            inv.invoice_number,
            inv.variable_symbol,
            inv.client.name,
            inv.client.ico or '',
            inv.issue_date.strftime('%d.%m.%Y'),
            inv.due_date.strftime('%d.%m.%Y'),
            inv.paid_date.strftime('%d.%m.%Y') if inv.paid_date else '',
            inv.subtotal,
            inv.vat_amount,
            inv.total,
            get_status_label(inv.status)
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col in [8, 9, 10]:  # Cisla
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
    
    # Sirka stlpcov
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 15
    
    # Ulozenie do pamati
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename=faktury_export_{date.today().strftime("%Y%m%d")}.xlsx'
        }
    )


@app.route('/invoices/export/xml')
@login_required
def invoices_export_xml():
    """Export faktur do XML (format pre uctovne systemy)"""
    invoices = Invoice.query.filter_by(user_id=current_user.id).order_by(Invoice.issue_date.desc()).all()
    supplier = Supplier.query.filter_by(user_id=current_user.id).first()
    
    xml_content = '<?xml version="1.0" encoding="UTF-8"?>\n'
    xml_content += '<Faktury xmlns="http://fakturask.sk/export" verzia="1.0">\n'
    xml_content += f'  <Export datum="{date.today().isoformat()}" />\n'
    
    if supplier:
        xml_content += '  <Dodavatel>\n'
        xml_content += f'    <Nazov>{supplier.name}</Nazov>\n'
        xml_content += f'    <ICO>{supplier.ico}</ICO>\n'
        xml_content += f'    <DIC>{supplier.dic or ""}</DIC>\n'
        xml_content += f'    <ICDPH>{supplier.ic_dph or ""}</ICDPH>\n'
        xml_content += f'    <Adresa>\n'
        xml_content += f'      <Ulica>{supplier.street}</Ulica>\n'
        xml_content += f'      <Mesto>{supplier.city}</Mesto>\n'
        xml_content += f'      <PSC>{supplier.zip_code}</PSC>\n'
        xml_content += f'    </Adresa>\n'
        xml_content += f'    <IBAN>{supplier.iban or ""}</IBAN>\n'
        xml_content += '  </Dodavatel>\n'
    
    xml_content += '  <ZoznamFaktur>\n'
    
    for inv in invoices:
        xml_content += f'    <Faktura id="{inv.id}">\n'
        xml_content += f'      <CisloFaktury>{inv.invoice_number}</CisloFaktury>\n'
        xml_content += f'      <VariabilnySymbol>{inv.variable_symbol}</VariabilnySymbol>\n'
        xml_content += f'      <DatumVystavenia>{inv.issue_date.isoformat()}</DatumVystavenia>\n'
        xml_content += f'      <DatumDodania>{inv.delivery_date.isoformat()}</DatumDodania>\n'
        xml_content += f'      <DatumSplatnosti>{inv.due_date.isoformat()}</DatumSplatnosti>\n'
        if inv.paid_date:
            xml_content += f'      <DatumUhrady>{inv.paid_date.isoformat()}</DatumUhrady>\n'
        xml_content += f'      <Stav>{inv.status}</Stav>\n'
        xml_content += f'      <FormaUhrady>{inv.payment_method}</FormaUhrady>\n'
        
        xml_content += f'      <Odberatel>\n'
        xml_content += f'        <Nazov>{inv.client.name}</Nazov>\n'
        xml_content += f'        <ICO>{inv.client.ico or ""}</ICO>\n'
        xml_content += f'        <DIC>{inv.client.dic or ""}</DIC>\n'
        xml_content += f'        <Adresa>\n'
        xml_content += f'          <Ulica>{inv.client.street}</Ulica>\n'
        xml_content += f'          <Mesto>{inv.client.city}</Mesto>\n'
        xml_content += f'          <PSC>{inv.client.zip_code}</PSC>\n'
        xml_content += f'        </Adresa>\n'
        xml_content += f'      </Odberatel>\n'
        
        xml_content += f'      <Polozky>\n'
        for item in inv.items:
            xml_content += f'        <Polozka>\n'
            xml_content += f'          <Popis>{item.description}</Popis>\n'
            xml_content += f'          <Mnozstvo>{item.quantity}</Mnozstvo>\n'
            xml_content += f'          <Jednotka>{item.unit}</Jednotka>\n'
            xml_content += f'          <JednotkovaCena>{item.unit_price:.2f}</JednotkovaCena>\n'
            xml_content += f'          <Spolu>{item.total:.2f}</Spolu>\n'
            xml_content += f'        </Polozka>\n'
        xml_content += f'      </Polozky>\n'
        
        xml_content += f'      <Sumy>\n'
        xml_content += f'        <ZakladDane>{inv.subtotal:.2f}</ZakladDane>\n'
        xml_content += f'        <SadzbaDPH>{inv.vat_rate:.0f}</SadzbaDPH>\n'
        xml_content += f'        <DPH>{inv.vat_amount:.2f}</DPH>\n'
        xml_content += f'        <Celkom>{inv.total:.2f}</Celkom>\n'
        xml_content += f'      </Sumy>\n'
        
        xml_content += f'    </Faktura>\n'
    
    xml_content += '  </ZoznamFaktur>\n'
    xml_content += '</Faktury>'
    
    return Response(
        xml_content,
        mimetype='application/xml',
        headers={
            'Content-Disposition': f'attachment; filename=faktury_export_{date.today().strftime("%Y%m%d")}.xml',
            'Content-Type': 'application/xml; charset=utf-8'
        }
    )


# ==============================================================================
# SPUSTENIE APLIKACIE
# ==============================================================================

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
